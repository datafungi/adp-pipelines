r"""Airflow hook for SharePoint Online, authenticated as an Entra ID app
registration with app-only permissions scoped via Sites.Selected.

Certificate-only, which is SharePoint's rule rather than a preference: it refuses
app-only tokens whose `appidacr` claim isn't "2" (certificate), so a client secret
401s however it's consented. Legacy ACS is not the way out -- it can't be scoped
with Sites.Selected at all. See docs/providers/sharepoint/service-principal-setup.md.

Connection (conn_type="sharepoint"), native fields only -- nothing in `extra`:
    login    -> application (client) ID
    host     -> tenant domain, e.g. "<tenant>.onmicrosoft.com"
    schema   -> certificate thumbprint
    password -> certificate private key (PEM, newlines escaped as \n)

`password` holds the key because Airflow masks and encrypts that field; the cost is
that it's single-line, hence the escaping. The key must be unencrypted -- no
passphrase is passed to MSAL. Field labels live in
provider_registration/adp_provider_reg/get_provider_info.py.

site_url is not part of the connection: Sites.Selected grants are per site, so it
comes from the DAG folder's config.toml (see plugins/utils/dag_config.py). Tokens
scope to the SharePoint resource, not Graph.

File paths are passed through as-is: SharePoint resolves them against the site, so
"Shared Documents/x.xlsx", "/Shared Documents/x.xlsx" and the full
"/sites/<site>/..." all reach the same file, and a path naming another site
resolves under this one and 404s rather than escaping it.
"""

from __future__ import annotations

import io
import re
from typing import TYPE_CHECKING, Any
from urllib.parse import urlsplit

from airflow.sdk.bases.hook import BaseHook
from cryptography.exceptions import UnsupportedAlgorithm
from cryptography.hazmat.primitives.serialization import load_pem_private_key
from office365.runtime.auth.entra.authentication_context import (
    AuthenticationContext as EntraAuthenticationContext,
)
from office365.sharepoint.client_context import ClientContext
from openpyxl import load_workbook

if TYPE_CHECKING:
    import pandas as pd
    from openpyxl.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

# pandas names a column "Unnamed: N" only when it consumed a header row whose cell was
# blank -- `header=None` yields integers instead. So the name is a reliable tell.
_UNNAMED = re.compile(r"^Unnamed: \d+$")

# Rows to scan under a suspect column before concluding it holds no formula. Bounds the
# cost on the pathological path; an all-empty column's formulas start at the first row.
_FORMULA_SCAN_ROWS = 200


def _load_private_key(private_key: str, conn_id: str) -> str:
    r"""Return the PEM with real line breaks, having checked it parses.

    Unescaping is unconditional: Airflow's `password` field is single-line and
    AIRFLOW_CONN_* is JSON, so a PEM can only arrive `\n`-escaped. PEM armour and
    base64 never contain a backslash, so `\n` is never key material.

    Parsing here is purely for the error message. MSAL fails deep inside its client
    assertion with `InvalidKeyError: Could not parse the provided public key`, which
    names neither the connection nor the cause, and is reached only after a token
    request -- so the same message covers a wrong key, an encrypted key, and a
    pasted certificate.
    """
    pem = private_key.replace("\\r\\n", "\n").replace("\\n", "\n")
    try:
        load_pem_private_key(pem.encode(), password=None)
    except TypeError as exc:
        raise ValueError(
            f"Connection '{conn_id}' has an encrypted 'private_key', which this hook "
            "cannot decrypt -- it passes no passphrase to MSAL. Re-issue the key "
            "unencrypted (openssl req ... -nodes)."
        ) from exc
    except (ValueError, UnsupportedAlgorithm) as exc:
        raise ValueError(
            f"Connection '{conn_id}' has a 'private_key' that is not a readable PEM "
            f"private key ({exc}). Check it is the private key (private-key.pem) and "
            "not the certificate, and that its newlines are escaped as \\n."
        ) from exc
    return pem


def _worksheet(workbook: Workbook, sheet_name: Any) -> Worksheet:
    """The sheet pandas would have read for `sheet_name`."""
    if isinstance(sheet_name, int):
        return workbook.worksheets[sheet_name]
    return workbook[sheet_name]


def _check_header(df: pd.DataFrame, file_path: str) -> None:
    """Raise if pandas took the wrong row as the header.

    `header=` is the fix, not the detection -- nobody passes it until the file breaks,
    and it breaks by luck. A title row leaves every column object dtype, which dies at
    XCom with `ArrowTypeError ... column Unnamed: 9` only if a datetime is present; an
    all-text file instead reaches Iceberg with columns named "Unnamed: 1".
    """
    columns = [str(column) for column in df.columns]
    unnamed = [column for column in columns if _UNNAMED.match(column)]
    if len(unnamed) * 2 <= len(columns):
        return

    raise ValueError(
        f"'{file_path}' parsed with {len(unnamed)} of {len(columns)} columns unnamed, "
        f"so the header row is wrong -- row 1 reads {columns[0]!r}. Pass header=<n>, "
        f"the 0-based index of the real header row, or check_header=False to override."
    )


def _formula_columns(
    raw: bytes, sheet_name: Any, suspects: list[str]
) -> dict[str, str]:
    """Map each of `suspects` that holds formulas to one sample formula.

    Finds the header by matching `suspects` against cell values, not by replaying
    skiprows/header/usecols, so it holds whatever the caller passed.
    """
    workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=False)
    try:
        rows = _worksheet(workbook, sheet_name).iter_rows(values_only=True)
        offsets: dict[str, int] = {}
        for row in rows:
            names = {str(cell): i for i, cell in enumerate(row) if cell is not None}
            offsets = {s: names[s] for s in suspects if s in names}
            if offsets:
                break
        if not offsets:
            return {}

        found: dict[str, str] = {}
        for _, row in zip(range(_FORMULA_SCAN_ROWS), rows):
            for column, offset in offsets.items():
                if column in found or offset >= len(row):
                    continue
                value = row[offset]
                if isinstance(value, str) and value.startswith("="):
                    found[column] = value
            if len(found) == len(offsets):
                break
        return found
    finally:
        workbook.close()


class SharePointHook(BaseHook):
    """Hook for authenticating to a SharePoint Online site."""

    conn_name_attr = "sharepoint_conn_id"
    default_conn_name = "sharepoint_default"
    conn_type = "sharepoint"
    hook_name = "SharePoint"

    def __init__(
        self, site_url: str, sharepoint_conn_id: str = default_conn_name
    ) -> None:
        super().__init__()
        self.site_url = site_url
        self.sharepoint_conn_id = sharepoint_conn_id
        self._client_context: ClientContext | None = None

    def get_conn(self) -> ClientContext:
        """Return an authenticated ClientContext for `self.site_url`, caching it."""
        if self._client_context is not None:
            return self._client_context

        conn = self.get_connection(self.sharepoint_conn_id)
        if not conn.login:
            raise ValueError(
                f"Connection '{self.sharepoint_conn_id}' is missing 'client_id'."
            )
        if not conn.host:
            raise ValueError(
                f"Connection '{self.sharepoint_conn_id}' is missing 'tenant'."
            )

        if not conn.schema:
            raise ValueError(
                f"Connection '{self.sharepoint_conn_id}' is missing 'thumbprint'."
            )
        if not conn.password:
            raise ValueError(
                f"Connection '{self.sharepoint_conn_id}' is missing 'private_key'."
            )

        scopes = [f"{self._site_resource()}/.default"]
        self.log.debug("Acquiring SharePoint token for scope: %s", scopes[0])

        entra_auth = EntraAuthenticationContext(
            tenant=conn.host, scopes=scopes
        ).with_certificate(
            conn.login,
            conn.schema,
            _load_private_key(conn.password, self.sharepoint_conn_id),
        )

        self._client_context = ClientContext(self.site_url).with_access_token(
            entra_auth.acquire_token
        )
        return self._client_context

    def _site_resource(self) -> str:
        """Resource identifier (scheme + host) the access token must be scoped to."""
        parts = urlsplit(self.site_url)
        return f"{parts.scheme}://{parts.netloc}"

    def download_file(self, file_path: str) -> bytes:
        """Return the bytes of `file_path`, resolved against the site.

        A missing file raises office365's ClientRequestException, which quotes the path
        SharePoint resolved -- more use than anything wrapped around it.
        """
        buffer = io.BytesIO()
        self.get_conn().web.get_file_by_server_relative_path(file_path).download(
            buffer
        ).execute_query()
        self.log.info("Downloaded %s (%d bytes)", file_path, buffer.tell())
        return buffer.getvalue()

    def read_excel(
        self,
        file_path: str,
        *,
        check_header: bool = True,
        check_formulas: bool = True,
        **pandas_kwargs: Any,
    ) -> pd.DataFrame | dict[str, pd.DataFrame]:
        """Read `file_path` into a DataFrame. `pandas_kwargs` go to `pd.read_excel`.

        Both checks are opt-out per call; see `_check_header` and
        `_warn_uncached_formulas`. `sheet_name=None` or a list makes pandas return a
        dict of frames, which the checks skip -- they read one sheet.
        """
        # Local: pandas costs ~200ms to import, which every DAG parse would pay for a
        # hook that is often only used to authenticate.
        import pandas as pd

        raw = self.download_file(file_path)
        # Annotated because `**pandas_kwargs` defeats read_excel's overloads, leaving
        # its return Any.
        df: pd.DataFrame | dict[str, pd.DataFrame] = pd.read_excel(
            io.BytesIO(raw), **pandas_kwargs
        )
        if not isinstance(df, pd.DataFrame):
            return df

        if check_header:
            _check_header(df, file_path)
        if check_formulas:
            self._warn_uncached_formulas(
                df, file_path, raw, pandas_kwargs.get("sheet_name", 0)
            )
        return df

    def _warn_uncached_formulas(
        self, df: pd.DataFrame, file_path: str, raw: bytes, sheet_name: Any
    ) -> None:
        """Warn for each empty column whose cells are really uncached formulas.

        Excel caches a formula's result on save; a library-written file never has, so
        the column arrives empty with correct dtype and nulls all the way into Iceberg
        and nothing else reports it. Keyed off an already-empty column so clean files
        skip the re-parse; openpyxl because only it sees formulas (calamine gives "").
        """
        suspects = [str(c) for c in df.columns if df[c].isna().all()]
        if not suspects:
            return

        for column, formula in _formula_columns(raw, sheet_name, suspects).items():
            self.log.warning(
                "Column '%s' in '%s' is entirely empty because its cells hold formulas "
                "with no cached value (e.g. %s). Excel caches results on save; files "
                "written by a library never do. Pass check_formulas=False to silence.",
                column,
                file_path,
                formula,
            )
