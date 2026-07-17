"""Tests for providers.sharepoint.hooks.sharepoint.SharePointHook."""

from __future__ import annotations

import io

import pytest
from adp_provider_reg.get_provider_info import get_provider_info
from cryptography.hazmat.primitives import serialization
from cryptography.hazmat.primitives.asymmetric import rsa
from openpyxl import Workbook

from providers.sharepoint.hooks.sharepoint import SharePointHook

SITE_URL = "https://contoso.sharepoint.com/sites/TalentAcquisition"
CLIENT_ID = "11111111-1111-1111-1111-111111111111"
TENANT_DOMAIN = "contoso.onmicrosoft.com"
THUMBPRINT = "AA:BB:CC"
EXPECTED_SCOPE = "https://contoso.sharepoint.com/.default"

_KEY = rsa.generate_private_key(public_exponent=65537, key_size=2048)


def _pem(encryption=serialization.NoEncryption()) -> str:
    return _KEY.private_bytes(
        serialization.Encoding.PEM, serialization.PrivateFormat.PKCS8, encryption
    ).decode()


# The hook validates the key, so these must be real PEMs. ESCAPED is the form the
# connection actually holds -- Airflow's single-line `password` field can't take
# literal newlines.
PRIVATE_KEY = _pem()
ESCAPED_PRIVATE_KEY = PRIVATE_KEY.replace("\n", "\\n")
ENCRYPTED_PRIVATE_KEY = _pem(serialization.BestAvailableEncryption(b"hunter2"))
PUBLIC_KEY = (
    _KEY.public_key()
    .public_bytes(
        serialization.Encoding.PEM, serialization.PublicFormat.SubjectPublicKeyInfo
    )
    .decode()
)


class FakeConnection:
    """Stands in for an Airflow Connection.

    No `extra_dejson` on purpose: the hook reads native fields only, so an
    AttributeError here is a real failure, not a fixture gap.
    """

    def __init__(
        self,
        login=CLIENT_ID,
        host=TENANT_DOMAIN,
        schema=THUMBPRINT,
        password=ESCAPED_PRIVATE_KEY,
    ):
        self.login = login
        self.host = host
        self.schema = schema
        self.password = password


def _patch_get_connection(mocker, **kwargs):
    conn = FakeConnection(**kwargs)
    return mocker.patch.object(SharePointHook, "get_connection", return_value=conn)


def _patch_auth_chain(mocker):
    """Patch the Entra AuthenticationContext + ClientContext bridge and return both mocks."""
    fake_entra_auth = mocker.Mock()
    fake_entra_auth.with_certificate.return_value = fake_entra_auth
    fake_entra_auth.with_client_secret.return_value = fake_entra_auth
    entra_auth_cls = mocker.patch(
        "providers.sharepoint.hooks.sharepoint.EntraAuthenticationContext",
        return_value=fake_entra_auth,
    )

    fake_ctx = mocker.Mock()
    fake_ctx.with_access_token.return_value = fake_ctx
    client_context_cls = mocker.patch(
        "providers.sharepoint.hooks.sharepoint.ClientContext", return_value=fake_ctx
    )

    return entra_auth_cls, fake_entra_auth, client_context_cls, fake_ctx


def test_hook_class_metadata():
    assert SharePointHook.conn_type == "sharepoint"
    assert SharePointHook.hook_name == "SharePoint"
    assert SharePointHook.default_conn_name == "sharepoint_default"
    assert SharePointHook.conn_name_attr == "sharepoint_conn_id"


def test_get_conn_authenticates_via_certificate(mocker):
    """with_client_secret must never be called: SharePoint refuses non-appidacr=2
    tokens, and an earlier revision regressed onto exactly that path."""
    get_connection = _patch_get_connection(mocker)
    entra_auth_cls, fake_entra_auth, client_context_cls, fake_ctx = _patch_auth_chain(
        mocker
    )

    hook = SharePointHook(site_url=SITE_URL, sharepoint_conn_id="sharepoint_default")
    result = hook.get_conn()

    entra_auth_cls.assert_called_once_with(
        tenant=TENANT_DOMAIN, scopes=[EXPECTED_SCOPE]
    )
    fake_entra_auth.with_certificate.assert_called_once_with(
        CLIENT_ID, THUMBPRINT, PRIVATE_KEY
    )
    fake_entra_auth.with_client_secret.assert_not_called()
    client_context_cls.assert_called_once_with(SITE_URL)
    fake_ctx.with_access_token.assert_called_once_with(fake_entra_auth.acquire_token)
    get_connection.assert_called_once_with("sharepoint_default")
    assert result is fake_ctx


def test_get_conn_caches_client_context(mocker):
    _patch_get_connection(mocker)
    _entra_auth_cls, _fake_entra_auth, client_context_cls, _fake_ctx = (
        _patch_auth_chain(mocker)
    )

    hook = SharePointHook(site_url=SITE_URL)
    first = hook.get_conn()
    second = hook.get_conn()

    assert first is second
    client_context_cls.assert_called_once()


@pytest.mark.parametrize("login", [None, ""])
def test_get_conn_missing_login_raises(mocker, login):
    _patch_get_connection(mocker, login=login)
    _patch_auth_chain(mocker)

    hook = SharePointHook(site_url=SITE_URL, sharepoint_conn_id="sharepoint_default")

    with pytest.raises(ValueError, match="client_id"):
        hook.get_conn()


@pytest.mark.parametrize("host", [None, ""])
def test_get_conn_missing_host_raises(mocker, host):
    _patch_get_connection(mocker, host=host)
    _patch_auth_chain(mocker)

    hook = SharePointHook(site_url=SITE_URL, sharepoint_conn_id="sharepoint_default")

    with pytest.raises(ValueError, match="'tenant'"):
        hook.get_conn()


@pytest.mark.parametrize("schema", [None, ""])
def test_get_conn_missing_schema_raises(mocker, schema):
    _patch_get_connection(mocker, schema=schema)
    _patch_auth_chain(mocker)

    hook = SharePointHook(site_url=SITE_URL, sharepoint_conn_id="sharepoint_default")

    with pytest.raises(ValueError, match="thumbprint"):
        hook.get_conn()


@pytest.mark.parametrize("password", [None, ""])
def test_get_conn_missing_password_raises(mocker, password):
    _patch_get_connection(mocker, password=password)
    _patch_auth_chain(mocker)

    hook = SharePointHook(site_url=SITE_URL, sharepoint_conn_id="sharepoint_default")

    with pytest.raises(ValueError, match="private_key"):
        hook.get_conn()


@pytest.mark.parametrize(
    "stored",
    [
        ESCAPED_PRIVATE_KEY,
        PRIVATE_KEY.replace("\n", "\\r\\n"),
        PRIVATE_KEY,
    ],
    ids=["escaped-lf", "escaped-crlf", "already-real-newlines"],
)
def test_get_conn_unescapes_private_key_newlines(mocker, stored):
    r"""MSAL can't parse a PEM without real line breaks, and the single-line
    `password` field can only receive them escaped. Real newlines pass through."""
    _patch_get_connection(mocker, password=stored)
    _entra_auth_cls, fake_entra_auth, _client_context_cls, _fake_ctx = (
        _patch_auth_chain(mocker)
    )

    SharePointHook(site_url=SITE_URL).get_conn()

    fake_entra_auth.with_certificate.assert_called_once_with(
        CLIENT_ID, THUMBPRINT, PRIVATE_KEY
    )


def test_get_conn_encrypted_private_key_raises(mocker):
    """A passphrase-protected key is a distinct mistake (skipping openssl's -nodes)
    from a malformed one, and the hook passes no passphrase to MSAL."""
    _patch_get_connection(mocker, password=ENCRYPTED_PRIVATE_KEY.replace("\n", "\\n"))
    _patch_auth_chain(mocker)

    hook = SharePointHook(site_url=SITE_URL, sharepoint_conn_id="sharepoint_default")

    with pytest.raises(ValueError, match="encrypted 'private_key'"):
        hook.get_conn()


def test_get_conn_certificate_pasted_as_private_key_raises(mocker):
    """Step 2 of the setup doc produces certificate.pem *and* private-key.pem, so
    pasting the wrong half is an easy mistake -- it must not reach MSAL."""
    _patch_get_connection(mocker, password=PUBLIC_KEY.replace("\n", "\\n"))
    _patch_auth_chain(mocker)

    hook = SharePointHook(site_url=SITE_URL, sharepoint_conn_id="sharepoint_default")

    with pytest.raises(ValueError, match="not a readable PEM private key"):
        hook.get_conn()


def test_get_conn_malformed_private_key_raises(mocker):
    """Without this the failure surfaces from inside MSAL as `InvalidKeyError: Could
    not parse the provided public key`, naming neither the connection nor the cause."""
    _patch_get_connection(mocker, password="-----BEGIN PRIVATE KEY-----\\nnope\\n")
    _patch_auth_chain(mocker)

    hook = SharePointHook(site_url=SITE_URL, sharepoint_conn_id="sharepoint_default")

    with pytest.raises(ValueError, match="not a readable PEM private key"):
        hook.get_conn()


def test_get_conn_private_key_error_names_the_connection(mocker):
    """The message must identify which connection is at fault -- a deployment has
    several, and MSAL's own error identifies none of them."""
    _patch_get_connection(mocker, password="garbage")
    _patch_auth_chain(mocker)

    hook = SharePointHook(site_url=SITE_URL, sharepoint_conn_id="sharepoint-prod-conn")

    with pytest.raises(ValueError, match="sharepoint-prod-conn"):
        hook.get_conn()


# --- provider_registration contract ---------------------------------------------

# Native Connection fields a form can expose. The hook reads other attributes that are
# no part of the schema the UI describes.
_NATIVE_FIELDS = frozenset(
    {
        "conn_id",
        "conn_type",
        "description",
        "extra",
        "host",
        "login",
        "password",
        "port",
        "schema",
    }
)


class RecordingConnection(FakeConnection):
    """FakeConnection that records which native fields were read.

    `__getattribute__`, not `__getattr__`: FakeConnection assigns the fields in
    `__init__`, so they exist and `__getattr__` would never fire. Unset native fields
    answer with a placeholder so a newly-read field fails the assert below rather than
    raising AttributeError, which would read as a fixture gap. Non-native attributes
    keep FakeConnection's behaviour of raising.
    """

    def __init__(self, **kwargs):
        object.__setattr__(self, "accessed", set())
        super().__init__(**kwargs)

    def __getattribute__(self, name):
        if name not in _NATIVE_FIELDS:
            return object.__getattribute__(self, name)

        object.__getattribute__(self, "accessed").add(name)
        try:
            return object.__getattribute__(self, name)
        except AttributeError:
            return "recorded"


def test_relabeled_fields_are_the_fields_the_hook_reads(mocker):
    """The registration and the hook must describe one schema.

    The registration is baked into the image, the hook git-synced, so they drift
    independently and nothing else compares them. Adding a `conn.port` read here would
    leave `port` hidden in the form and the connection unconfigurable.

    Scoped to get_conn() because that is where a hook consumes its Connection -- the
    other methods take paths, not credentials.
    """
    conn = RecordingConnection()
    mocker.patch.object(SharePointHook, "get_connection", return_value=conn)
    _patch_auth_chain(mocker)

    SharePointHook(site_url=SITE_URL).get_conn()

    entry = next(
        e
        for e in get_provider_info()["connection-types"]
        if e["connection-type"] == SharePointHook.conn_type
    )
    assert conn.accessed == set(entry["ui-field-behaviour"]["relabeling"])


# --- download_file / read_excel -------------------------------------------------
#
# Real xlsx built by openpyxl, mirroring the seed data's deliberate defects (see
# scripts/generate_seeds.py). The checks catch what a real parser does to a real
# file, so an imitation fixture would test nothing.

FILE_PATH = "Shared Documents/RecruitingDocuments/Requisitions/x.xlsx"


def _xlsx(
    rows: list[list], sheet_name: str = "Sheet1", extra: dict | None = None
) -> bytes:
    workbook = Workbook()
    workbook.active.title = sheet_name
    for row in rows:
        workbook.active.append(row)
    for name, extra_rows in (extra or {}).items():
        sheet = workbook.create_sheet(name)
        for row in extra_rows:
            sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


CLEAN_XLSX = _xlsx(
    [["req_id", "title", "status", "days_open"], ["R1", "Engineer", "Open", 10]]
)
# openpyxl writes the formula with no cached result, which is why it reads back empty.
FORMULA_XLSX = _xlsx(
    [
        ["req_id", "open_date", "days_open"],
        ["R1", "2026-01-01", '=DATEDIF(B2,TODAY(),"d")'],
        ["R2", "2026-01-02", '=DATEDIF(B3,TODAY(),"d")'],
    ]
)
# Title row + blank row above the real header, as in candidates weeks 3 and 15.
TITLE_XLSX = _xlsx(
    [
        ["Nimbus Robotics — Open Requisitions — CONFIDENTIAL"],
        [],
        ["req_id", "title", "status", "days_open"],
        ["R1", "Engineer", "Open", 10],
    ]
)
EMPTY_COLUMN_XLSX = _xlsx([["req_id", "notes"], ["R1", None], ["R2", None]])


def _patch_download(mocker, raw: bytes = CLEAN_XLSX):
    return mocker.patch.object(SharePointHook, "download_file", return_value=raw)


def _patch_log(mocker):
    """Stand in for the hook's logger.

    Not caplog: Airflow 3 routes `self.log` through structlog, so `caplog.text` stays
    empty however much is logged -- quietly passing any assert-not-logged.
    """
    return mocker.patch.object(SharePointHook, "log")


def test_download_file_passes_the_path_through_untouched(mocker):
    """SharePoint resolves paths against the site, so site-relative, "/"-prefixed and
    full server-relative forms all land on the same file. Rewriting could only break
    that."""
    fake_file = mocker.Mock()
    fake_file.download.side_effect = lambda buffer: (
        buffer.write(b"xlsx-bytes"),
        fake_file,
    )[1]
    fake_ctx = mocker.Mock()
    fake_ctx.web.get_file_by_server_relative_path.return_value = fake_file
    mocker.patch.object(SharePointHook, "get_conn", return_value=fake_ctx)

    result = SharePointHook(site_url=SITE_URL).download_file(FILE_PATH)

    fake_ctx.web.get_file_by_server_relative_path.assert_called_once_with(FILE_PATH)
    assert result == b"xlsx-bytes"


def test_read_excel_returns_a_dataframe(mocker):
    download_file = _patch_download(mocker)

    df = SharePointHook(site_url=SITE_URL).read_excel(FILE_PATH)

    download_file.assert_called_once_with(FILE_PATH)
    assert list(df.columns) == ["req_id", "title", "status", "days_open"]
    assert df.shape == (1, 4)


def test_read_excel_forwards_pandas_kwargs(mocker):
    _patch_download(mocker, TITLE_XLSX)

    df = SharePointHook(site_url=SITE_URL).read_excel(FILE_PATH, header=2)

    assert list(df.columns) == ["req_id", "title", "status", "days_open"]


def test_read_excel_reads_a_named_sheet(mocker):
    _patch_download(
        mocker,
        _xlsx(
            [["req_id"], ["R1"]],
            sheet_name="Requisitions",
            extra={"Departments": [["department_id"], ["D1"]]},
        ),
    )

    df = SharePointHook(site_url=SITE_URL).read_excel(
        FILE_PATH, sheet_name="Departments"
    )

    assert list(df.columns) == ["department_id"]


def test_read_excel_multiple_sheets_returns_dict_and_skips_checks(mocker):
    """The checks read one sheet, so they must step aside rather than explode on a
    dict."""
    _patch_download(
        mocker, _xlsx([["req_id"], ["R1"]], extra={"Departments": [["d_id"], ["D1"]]})
    )

    result = SharePointHook(site_url=SITE_URL).read_excel(FILE_PATH, sheet_name=None)

    assert isinstance(result, dict)
    assert set(result) == {"Sheet1", "Departments"}


# --- check_header ---------------------------------------------------------------


def test_read_excel_raises_when_a_title_row_became_the_header(mocker):
    """Whether a title row dies at XCom or reaches Iceberg depends on the column types,
    so it can't be left to chance."""
    _patch_download(mocker, TITLE_XLSX)

    with pytest.raises(ValueError, match="3 of 4 columns unnamed"):
        SharePointHook(site_url=SITE_URL).read_excel(FILE_PATH)


def test_header_error_quotes_the_offending_row_and_names_the_file(mocker):
    """The quoted row is the whole diagnosis, without opening the file."""
    _patch_download(mocker, TITLE_XLSX)

    with pytest.raises(ValueError) as excinfo:
        SharePointHook(site_url=SITE_URL).read_excel(FILE_PATH)

    message = str(excinfo.value)
    assert FILE_PATH in message
    assert "Nimbus Robotics — Open Requisitions — CONFIDENTIAL" in message
    assert "header=<n>" in message


def test_check_header_false_returns_the_frame_unchecked(mocker):
    """It hands back the mangled frame, real header sitting in the data -- which is the
    point of having to ask for it."""
    _patch_download(mocker, TITLE_XLSX)

    df = SharePointHook(site_url=SITE_URL).read_excel(FILE_PATH, check_header=False)

    assert df.shape == (3, 4)
    assert list(df.iloc[1]) == ["req_id", "title", "status", "days_open"]


def test_read_excel_does_not_flag_a_healthy_header(mocker):
    _patch_download(mocker, CLEAN_XLSX)

    SharePointHook(site_url=SITE_URL).read_excel(FILE_PATH)  # must not raise


# --- check_formulas -------------------------------------------------------------


def test_read_excel_warns_when_a_column_is_empty_because_of_formulas(mocker):
    """The dtype and nulls stay correct all the way into Iceberg, so nothing else
    reports this."""
    _patch_download(mocker, FORMULA_XLSX)
    log = _patch_log(mocker)

    df = SharePointHook(site_url=SITE_URL).read_excel(FILE_PATH)

    assert df["days_open"].isna().all()
    log.warning.assert_called_once()
    _, column, path, formula = log.warning.call_args.args
    assert column == "days_open"
    assert path == FILE_PATH
    assert formula == '=DATEDIF(B2,TODAY(),"d")'


def test_formula_check_is_skipped_when_no_column_is_empty(mocker):
    """The re-parse is the whole cost of this check, so a clean file must not pay it."""
    _patch_download(mocker, CLEAN_XLSX)
    load_workbook = mocker.patch("providers.sharepoint.hooks.sharepoint.load_workbook")

    SharePointHook(site_url=SITE_URL).read_excel(FILE_PATH)

    load_workbook.assert_not_called()


def test_no_warning_for_a_column_that_is_simply_empty(mocker):
    """Only worth reporting if formulas explain it; otherwise the file really is like
    that."""
    _patch_download(mocker, EMPTY_COLUMN_XLSX)
    log = _patch_log(mocker)

    df = SharePointHook(site_url=SITE_URL).read_excel(FILE_PATH)

    assert df["notes"].isna().all()
    log.warning.assert_not_called()


def test_check_formulas_false_silences_the_warning(mocker):
    _patch_download(mocker, FORMULA_XLSX)
    log = _patch_log(mocker)

    SharePointHook(site_url=SITE_URL).read_excel(FILE_PATH, check_formulas=False)

    log.warning.assert_not_called()


def test_formula_check_survives_a_header_offset(mocker):
    """Columns are found by name, not by replaying skiprows/header arithmetic."""
    _patch_download(
        mocker,
        _xlsx(
            [
                ["Nimbus Robotics — CONFIDENTIAL"],
                [],
                ["req_id", "open_date", "days_open"],
                ["R1", "2026-01-01", '=DATEDIF(B4,TODAY(),"d")'],
            ]
        ),
    )
    log = _patch_log(mocker)

    SharePointHook(site_url=SITE_URL).read_excel(FILE_PATH, header=2)

    assert log.warning.call_args.args[1] == "days_open"


def test_formula_check_reads_the_sheet_pandas_read(mocker):
    """Sheet 0 here is clean, so reading it instead would report nothing."""
    _patch_download(
        mocker,
        _xlsx(
            [["req_id", "days_open"], ["R1", 10]],
            sheet_name="Requisitions",
            extra={
                "Archive": [["req_id", "days_open"], ["R9", '=DATEDIF(A2,TODAY(),"d")']]
            },
        ),
    )
    log = _patch_log(mocker)

    SharePointHook(site_url=SITE_URL).read_excel(FILE_PATH, sheet_name="Archive")

    assert log.warning.call_args.args[1] == "days_open"
